from datetime import datetime

from flask import Blueprint, request, jsonify, send_file, render_template_string
from app.extensions import db
from app.models.key_person import KeyPerson
from app.models.person_contact import PersonContact
from app.models.person_case import PersonCase
from app.models.person_track import PersonTrack
from app.models.person_alert import PersonAlert
from app.models.visit_task import VisitTask
from app.models.visit_record import VisitRecord
from app.models.risk_assessment import RiskAssessment
from app.models.lost_contact_track import LostContactTrack
from app.api.auth import login_required, admin_required
from app.utils import log_operation

import_bp = Blueprint('import_export', __name__)


@import_bp.route('/export/persons', methods=['GET'])
@login_required
def export_persons():
    persons = KeyPerson.query.order_by(KeyPerson.created_at.desc()).all()
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '重点人员'

    headers = ['姓名', '性别', '身份证号', '出生日期', '手机号', '户籍地址', '现住地址',
               '学历', '就业状况', '工作单位', '政治面貌', '民族', '婚姻状况', '户籍类型',
               '人员类别', '风险等级', '管控状态', '主要事由']
    ws.append(headers)

    for p in persons:
        ws.append([
            p.name,
            '男' if p.gender == 'M' else '女' if p.gender == 'F' else '',
            p.id_card,
            p.birth_date.isoformat() if p.birth_date else '',
            p.phone or '',
            p.address or '',
            p.current_address or '',
            p.education or '',
            p.employment_status or '',
            p.employer or '',
            p.political_status or '',
            p.ethnicity or '',
            p.marital_status or '',
            p.household_type or '',
            p.category.category_name if p.category else '',
            {'high': '高风险', 'medium': '中风险', 'low': '低风险'}.get(p.risk_level, p.risk_level),
            {'monitored': '管控中', 'removed': '已撤销', 'archived': '已归档',
             'lost': '失联', 'missing': '下落不明'}.get(p.control_status, p.control_status),
            p.case_description or '',
        ])

    filename = f'重点人员导出_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    filepath = f'/tmp/{filename}'
    wb.save(filepath)

    log_operation('EXPORT', 'key_person', entity_name=f'导出{len(persons)}条记录')
    db.session.commit()

    return send_file(filepath, as_attachment=True, download_name=filename)


@import_bp.route('/import/persons', methods=['POST'])
@admin_required
def import_persons():
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '未选择文件'}), 400

    file = request.files['file']
    import openpyxl
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    success = 0
    errors = []

    gender_map = {'男': 'M', '女': 'F'}
    risk_map = {'高风险': 'high', '中风险': 'medium', '低风险': 'low'}
    status_map = {'管控中': 'monitored', '已撤销': 'removed', '已归档': 'archived',
                  '失联': 'lost', '下落不明': 'missing'}

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            name = str(row[0]).strip() if row[0] else ''
            id_card = str(row[2]).strip() if row[2] else ''
            if not name or not id_card:
                errors.append(f'第{i}行: 姓名和身份证号不能为空')
                continue

            existing = KeyPerson.query.filter_by(id_card=id_card).first()
            if existing:
                errors.append(f'第{i}行: 身份证号{id_card}已存在')
                continue

            person = KeyPerson(
                name=name,
                gender=gender_map.get(str(row[1]).strip()) if row[1] else None,
                id_card=id_card,
                birth_date=str(row[3]).strip() if row[3] else None,
                phone=str(row[4]).strip() if row[4] else '',
                address=str(row[5]).strip() if row[5] else '',
                current_address=str(row[6]).strip() if row[6] else '',
                education=str(row[7]).strip() if row[7] else '',
                employment_status=str(row[8]).strip() if row[8] else '',
                employer=str(row[9]).strip() if row[9] else '',
                political_status=str(row[10]).strip() if row[10] else '',
                ethnicity=str(row[11]).strip() if row[11] else '',
                marital_status=str(row[12]).strip() if row[12] else '',
                household_type=str(row[13]).strip() if row[13] else '',
                risk_level=risk_map.get(str(row[15]).strip(), 'medium') if row[15] else 'medium',
                control_status=status_map.get(str(row[16]).strip(), 'monitored') if row[16] else 'monitored',
                case_description=str(row[17]).strip() if row[17] else '',
                created_by=request.current_user['user_id'],
            )
            db.session.add(person)
            success += 1
        except Exception as e:
            errors.append(f'第{i}行: {str(e)}')

    db.session.commit()
    log_operation('IMPORT', 'key_person', entity_name=f'导入{success}条,失败{len(errors)}条')
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': f'导入完成: 成功{success}条, 失败{len(errors)}条',
        'data': {'success': success, 'errors': errors},
    })


@import_bp.route('/archive/<int:person_id>', methods=['GET'])
@login_required
def get_person_archive(person_id):
    person = KeyPerson.query.get(person_id)
    if not person:
        return jsonify({'code': 404, 'message': '人员不存在'}), 404

    contacts = PersonContact.query.filter_by(person_id=person_id).order_by(PersonContact.created_at.desc()).all()
    cases = PersonCase.query.filter_by(person_id=person_id).order_by(PersonCase.created_at.desc()).all()
    tracks = PersonTrack.query.filter_by(person_id=person_id).order_by(PersonTrack.track_time.desc()).all()
    alerts = PersonAlert.query.filter_by(person_id=person_id).order_by(PersonAlert.alert_time.desc()).all()
    visit_tasks = VisitTask.query.filter_by(person_id=person_id).order_by(VisitTask.created_at.desc()).all()
    visit_records = VisitRecord.query.filter_by(person_id=person_id).order_by(VisitRecord.visit_time.desc()).all()
    risk_assessments = RiskAssessment.query.filter_by(person_id=person_id).order_by(RiskAssessment.created_at.desc()).all()
    lost_contact_tracks = LostContactTrack.query.filter_by(person_id=person_id).order_by(LostContactTrack.created_at.desc()).all()

    data = {
        'person': person.to_dict(),
        'contacts': [c.to_dict() for c in contacts],
        'cases': [c.to_dict() for c in cases],
        'tracks': [t.to_dict() for t in tracks],
        'alerts': [a.to_dict() for a in alerts],
        'visit_tasks': [t.to_dict() for t in visit_tasks],
        'visit_records': [r.to_dict() for r in visit_records],
        'risk_assessments': [r.to_dict() for r in risk_assessments],
        'lost_contact_tracks': [l.to_dict() for l in lost_contact_tracks],
    }

    log_operation('PRINT', 'key_person', person_id, person.name)
    db.session.commit()

    return jsonify({'code': 200, 'data': data})
